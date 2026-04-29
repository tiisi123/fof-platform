"""
报表服务 - 统一报表数据汇总与导出
"""
import io
from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models.manager import Manager
from app.models.product import Product
from app.models.nav import NavData
from app.models.portfolio import Portfolio, PortfolioComponent
from app.models.project import Project
from app.services import performance_service


class ReportService:
    """报表服务"""
    
    # 样式定义
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, db: Session):
        self.db = db
    
    def _get_performance_metrics(self, product_id: int, period: str) -> dict:
        """获取业绩指标"""
        try:
            end_date = date.today()
            start_date = performance_service.get_period_start_date(end_date, period)
            nav_series = performance_service.get_nav_series(self.db, product_id, start_date, end_date)
            
            if len(nav_series) < 2:
                return {}
            
            daily_returns = performance_service.calculate_daily_returns(nav_series)
            return_data = performance_service.calculate_return(nav_series)
            vol_data = performance_service.calculate_volatility(daily_returns)
            dd_data = performance_service.calculate_max_drawdown(nav_series)
            
            result = {
                'total_return': return_data.get('total_return'),
                'annualized_return': return_data.get('annualized_return'),
                'annualized_volatility': vol_data.get('annualized_volatility'),
                'max_drawdown': dd_data.get('max_drawdown'),
            }
            
            # 计算夏普比率
            if result['annualized_return'] and result['annualized_volatility'] and result['annualized_volatility'] > 0:
                risk_free_rate = 0.02
                result['sharpe_ratio'] = (result['annualized_return'] - risk_free_rate) / result['annualized_volatility']
            else:
                result['sharpe_ratio'] = None
            
            # 计算卡玛比率
            if result['annualized_return'] and result['max_drawdown'] and result['max_drawdown'] > 0:
                result['calmar_ratio'] = result['annualized_return'] / result['max_drawdown']
            else:
                result['calmar_ratio'] = None
            
            return result
        except:
            return {}
    
    def _apply_header_style(self, ws, row_num: int, col_count: int):
        """应用表头样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
    
    def _apply_cell_style(self, ws, start_row: int, end_row: int, col_count: int, number_cols: List[int] = None):
        """应用单元格样式"""
        number_cols = number_cols or []
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                if col in number_cols:
                    cell.alignment = self.NUMBER_ALIGNMENT
                else:
                    cell.alignment = self.CELL_ALIGNMENT
    
    def _auto_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        # 中文字符算2个宽度
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ========== 管理人汇总报表 ==========
    
    def generate_manager_report(
        self,
        pool_category: Optional[str] = None,
        primary_strategy: Optional[str] = None
    ) -> io.BytesIO:
        """
        生成管理人汇总报表
        
        包含：管理人基本信息、旗下产品数量、最新业绩指标
        """
        # 查询管理人
        query = self.db.query(Manager).filter(Manager.is_deleted == False)
        if pool_category:
            query = query.filter(Manager.pool_category == pool_category)
        if primary_strategy:
            query = query.filter(Manager.primary_strategy == primary_strategy)
        managers = query.all()
        
        # 策略映射
        strategy_map = {
            "equity_long": "股票多头", "quant_neutral": "量化中性", "cta": "CTA",
            "arbitrage": "套利", "multi_strategy": "多策略", "bond": "债券", "other": "其他"
        }
        pool_map = {
            "invested": "在投池", "key_tracking": "重点跟踪池", "observation": "观察池",
            "eliminated": "淘汰池", "contacted": "已看过"
        }
        
        # 准备数据
        data = []
        for mgr in managers:
            # 获取旗下产品数量
            product_count = self.db.query(func.count(Product.id)).filter(
                Product.manager_id == mgr.id
            ).scalar()
            
            # 获取代表产品最新业绩（如果有）
            representative = self.db.query(Product).filter(
                Product.manager_id == mgr.id,
                Product.is_representative == True
            ).first()
            
            perf_data = {}
            if representative:
                try:
                    metrics = self._get_performance_metrics(
                        representative.id, "1y"
                    )
                    perf_data = {
                        "代表产品": representative.product_name,
                        "近1年收益": f"{metrics.get('total_return', 0) * 100:.2f}%" if metrics.get('total_return') else "-",
                        "年化波动": f"{metrics.get('annualized_volatility', 0) * 100:.2f}%" if metrics.get('annualized_volatility') else "-",
                        "最大回撤": f"{metrics.get('max_drawdown', 0) * 100:.2f}%" if metrics.get('max_drawdown') else "-",
                        "夏普比率": f"{metrics.get('sharpe_ratio', 0):.2f}" if metrics.get('sharpe_ratio') else "-",
                    }
                except:
                    pass
            
            row = {
                "管理人编号": mgr.manager_code or "",
                "管理人名称": mgr.manager_name,
                "管理人简称": mgr.short_name or "",
                "一级策略": strategy_map.get(mgr.primary_strategy, mgr.primary_strategy or ""),
                "跟踪池分类": pool_map.get(mgr.pool_category, mgr.pool_category or ""),
                "管理规模": mgr.aum_range or "",
                "产品数量": product_count,
                "协会备案号": mgr.registration_no or "",
                "成立日期": str(mgr.established_date) if mgr.established_date else "",
                **perf_data
            }
            data.append(row)
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "管理人汇总"
        
        # 写入数据
        df = pd.DataFrame(data)
        if df.empty:
            df = pd.DataFrame(columns=["管理人编号", "管理人名称", "管理人简称", "一级策略", "跟踪池分类"])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 应用样式
        self._apply_header_style(ws, 1, len(df.columns))
        if len(data) > 0:
            self._apply_cell_style(ws, 2, len(data) + 1, len(df.columns), [7])
        self._auto_column_width(ws)
        
        # 保存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ========== 跟踪池报表 ==========
    
    def generate_pool_report(self, pool_category: Optional[str] = None) -> io.BytesIO:
        """
        生成跟踪池报表
        
        包含：各分类管理人及其代表产品业绩
        """
        pool_map = {
            "invested": "在投池", "key_tracking": "重点跟踪池", "observation": "观察池",
            "eliminated": "淘汰池", "contacted": "已看过"
        }
        strategy_map = {
            "equity_long": "股票多头", "quant_neutral": "量化中性", "cta": "CTA",
            "arbitrage": "套利", "multi_strategy": "多策略", "bond": "债券", "other": "其他"
        }
        
        wb = Workbook()
        first_sheet = True
        
        pools = [pool_category] if pool_category else list(pool_map.keys())
        
        for pool in pools:
            pool_name = pool_map.get(pool, pool)
            
            # 查询该分类的管理人
            managers = self.db.query(Manager).filter(
                Manager.is_deleted == False,
                Manager.pool_category == pool
            ).all()
            
            data = []
            for mgr in managers:
                # 获取代表产品
                representative = self.db.query(Product).filter(
                    Product.manager_id == mgr.id,
                    Product.is_representative == True
                ).first()
                
                row = {
                    "管理人名称": mgr.manager_name,
                    "策略": strategy_map.get(mgr.primary_strategy, ""),
                    "代表产品": representative.product_name if representative else "-",
                }
                
                # 获取多周期业绩
                if representative:
                    for period, label in [("1m", "近1月"), ("3m", "近3月"), ("6m", "近6月"), ("1y", "近1年")]:
                        try:
                            metrics = self._get_performance_metrics(
                                representative.id, period
                            )
                            ret = metrics.get('total_return')
                            row[f"{label}收益"] = f"{ret * 100:.2f}%" if ret is not None else "-"
                        except:
                            row[f"{label}收益"] = "-"
                    
                    # 获取最大回撤和夏普
                    try:
                        metrics = self._get_performance_metrics(
                            representative.id, "1y"
                        )
                        row["最大回撤"] = f"{metrics.get('max_drawdown', 0) * 100:.2f}%" if metrics.get('max_drawdown') else "-"
                        row["夏普比率"] = f"{metrics.get('sharpe_ratio', 0):.2f}" if metrics.get('sharpe_ratio') else "-"
                    except:
                        row["最大回撤"] = "-"
                        row["夏普比率"] = "-"
                else:
                    for label in ["近1月收益", "近3月收益", "近6月收益", "近1年收益", "最大回撤", "夏普比率"]:
                        row[label] = "-"
                
                data.append(row)
            
            # 创建工作表
            if first_sheet:
                ws = wb.active
                ws.title = pool_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=pool_name)
            
            df = pd.DataFrame(data)
            if df.empty:
                df = pd.DataFrame(columns=["管理人名称", "策略", "代表产品", "近1月收益", "近3月收益", "近6月收益", "近1年收益", "最大回撤", "夏普比率"])
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            self._apply_header_style(ws, 1, len(df.columns))
            if len(data) > 0:
                self._apply_cell_style(ws, 2, len(data) + 1, len(df.columns), list(range(4, 10)))
            self._auto_column_width(ws)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ========== 产品业绩报表 ==========
    
    def generate_product_report(
        self,
        strategy_type: Optional[str] = None,
        period: str = "1y"
    ) -> io.BytesIO:
        """
        生成产品业绩报表
        
        包含：产品基本信息、多周期业绩指标
        """
        strategy_map = {
            "equity_long": "股票多头", "quant_neutral": "量化中性", "cta": "CTA",
            "arbitrage": "套利", "multi_strategy": "多策略", "bond": "债券", "other": "其他"
        }
        
        # 查询产品
        query = self.db.query(Product)
        if strategy_type:
            query = query.filter(Product.strategy_type == strategy_type)
        products = query.all()
        
        data = []
        for prod in products:
            # 获取管理人
            manager = self.db.query(Manager).filter(Manager.id == prod.manager_id).first()
            
            row = {
                "产品代码": prod.product_code or "",
                "产品名称": prod.product_name,
                "管理人": manager.manager_name if manager else "",
                "策略类型": strategy_map.get(prod.strategy_type, prod.strategy_type or ""),
                "成立日期": str(prod.inception_date) if prod.inception_date else "",
            }
            
            # 获取最新净值
            latest_nav = self.db.query(NavData).filter(
                NavData.product_id == prod.id
            ).order_by(NavData.nav_date.desc()).first()
            
            if latest_nav:
                row["最新净值日期"] = str(latest_nav.nav_date)
                row["单位净值"] = f"{latest_nav.unit_nav:.4f}"
                row["累计净值"] = f"{latest_nav.cumulative_nav:.4f}" if latest_nav.cumulative_nav else "-"
            else:
                row["最新净值日期"] = "-"
                row["单位净值"] = "-"
                row["累计净值"] = "-"
            
            # 获取业绩指标
            for p, label in [("1m", "近1月"), ("3m", "近3月"), ("6m", "近6月"), ("1y", "近1年"), ("ytd", "今年以来")]:
                try:
                    metrics = self._get_performance_metrics(prod.id, p)
                    ret = metrics.get('total_return')
                    row[f"{label}收益"] = f"{ret * 100:.2f}%" if ret is not None else "-"
                except:
                    row[f"{label}收益"] = "-"
            
            # 风险指标
            try:
                metrics = self._get_performance_metrics(prod.id, period)
                row["年化波动"] = f"{metrics.get('annualized_volatility', 0) * 100:.2f}%" if metrics.get('annualized_volatility') else "-"
                row["最大回撤"] = f"{metrics.get('max_drawdown', 0) * 100:.2f}%" if metrics.get('max_drawdown') else "-"
                row["夏普比率"] = f"{metrics.get('sharpe_ratio', 0):.2f}" if metrics.get('sharpe_ratio') else "-"
                row["卡玛比率"] = f"{metrics.get('calmar_ratio', 0):.2f}" if metrics.get('calmar_ratio') else "-"
            except:
                row["年化波动"] = "-"
                row["最大回撤"] = "-"
                row["夏普比率"] = "-"
                row["卡玛比率"] = "-"
            
            data.append(row)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "产品业绩"
        
        df = pd.DataFrame(data)
        if df.empty:
            df = pd.DataFrame(columns=["产品代码", "产品名称", "管理人", "策略类型"])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws, 1, len(df.columns))
        if len(data) > 0:
            self._apply_cell_style(ws, 2, len(data) + 1, len(df.columns))
        self._auto_column_width(ws)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ========== 组合报表 ==========
    
    def generate_portfolio_report(self, portfolio_id: Optional[int] = None) -> io.BytesIO:
        """
        生成组合报表
        
        包含：组合概览、成分明细、业绩指标
        """
        if portfolio_id:
            portfolios = [self.db.query(Portfolio).filter(Portfolio.id == portfolio_id).first()]
            portfolios = [p for p in portfolios if p]
        else:
            portfolios = self.db.query(Portfolio).all()
        
        wb = Workbook()
        
        # 概览表
        ws_overview = wb.active
        ws_overview.title = "组合概览"
        
        overview_data = []
        for pf in portfolios:
            components = self.db.query(PortfolioComponent).filter(
                PortfolioComponent.portfolio_id == pf.id
            ).all()
            
            row = {
                "组合名称": pf.name,
                "组合描述": pf.description or "",
                "成分数量": len(components),
                "创建时间": str(pf.created_at.date()) if pf.created_at else "",
            }
            
            # 组合业绩（需要专门的组合净值计算，这里简化处理）
            overview_data.append(row)
        
        df = pd.DataFrame(overview_data)
        if df.empty:
            df = pd.DataFrame(columns=["组合名称", "组合描述", "成分数量", "创建时间"])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_overview.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws_overview, 1, len(df.columns))
        self._auto_column_width(ws_overview)
        
        # 成分明细表
        ws_components = wb.create_sheet(title="成分明细")
        
        comp_data = []
        for pf in portfolios:
            components = self.db.query(PortfolioComponent).filter(
                PortfolioComponent.portfolio_id == pf.id
            ).all()
            
            for comp in components:
                product = self.db.query(Product).filter(Product.id == comp.product_id).first()
                if product:
                    manager = self.db.query(Manager).filter(Manager.id == product.manager_id).first()
                    comp_data.append({
                        "组合名称": pf.name,
                        "产品代码": product.product_code or "",
                        "产品名称": product.product_name,
                        "管理人": manager.manager_name if manager else "",
                        "权重": f"{comp.weight * 100:.2f}%",
                    })
        
        df_comp = pd.DataFrame(comp_data)
        if df_comp.empty:
            df_comp = pd.DataFrame(columns=["组合名称", "产品代码", "产品名称", "管理人", "权重"])
        
        for r_idx, row in enumerate(dataframe_to_rows(df_comp, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_components.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws_components, 1, len(df_comp.columns))
        self._auto_column_width(ws_components)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ========== 一级项目报表 ==========
    
    def generate_project_report(self, stage: Optional[str] = None) -> io.BytesIO:
        """
        生成一级项目汇总报表
        """
        stage_map = {
            "sourcing": "Sourcing", "initial_screening": "初筛", "due_diligence": "尽调",
            "investment_decision": "投决", "post_investment": "投后", "exit": "退出"
        }
        
        query = self.db.query(Project)
        if stage:
            query = query.filter(Project.stage == stage)
        projects = query.all()
        
        data = []
        for proj in projects:
            data.append({
                "项目名称": proj.project_name,
                "项目编号": proj.project_code or "",
                "行业": proj.industry or "",
                "当前阶段": stage_map.get(proj.stage, proj.stage or ""),
                "投资金额(万)": proj.investment_amount or "",
                "估值(万)": proj.valuation or "",
                "联系人": proj.contact_person or "",
                "联系方式": proj.contact_info or "",
                "来源": proj.source or "",
                "创建时间": str(proj.created_at.date()) if proj.created_at else "",
                "备注": proj.remark or "",
            })
        
        wb = Workbook()
        ws = wb.active
        ws.title = "一级项目"
        
        df = pd.DataFrame(data)
        if df.empty:
            df = pd.DataFrame(columns=["项目名称", "项目编号", "行业", "当前阶段", "投资金额(万)", "估值(万)"])
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws, 1, len(df.columns))
        if len(data) > 0:
            self._apply_cell_style(ws, 2, len(data) + 1, len(df.columns))
        self._auto_column_width(ws)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ========== 报表预览 ==========

    def get_preview_data(self, report_key: str, **kwargs) -> dict:
        """获取报表预览数据（返回JSON）"""
        preview_map = {
            'manager': self._preview_manager,
            'pool': self._preview_manager,  # 复用manager逻辑
            'product': self._preview_product,
            'portfolio': self._preview_portfolio,
            'project': self._preview_project,
        }
        func = preview_map.get(report_key)
        if not func:
            return {"columns": [], "rows": [], "total": 0, "error": f"未知报表类型: {report_key}"}
        return func(**kwargs)

    def _preview_manager(self, **kwargs) -> dict:
        pool_category = kwargs.get('pool_category')
        primary_strategy = kwargs.get('primary_strategy')
        strategy_map = {
            "equity_long": "股票多头", "quant_neutral": "量化中性", "cta": "CTA",
            "arbitrage": "套利", "multi_strategy": "多策略", "bond": "债券", "other": "其他"
        }
        pool_map = {
            "invested": "在投池", "key_tracking": "重点跟踪池", "observation": "观察池",
            "eliminated": "淘汰池", "contacted": "已看过"
        }
        query = self.db.query(Manager).filter(Manager.is_deleted == False)
        if pool_category:
            query = query.filter(Manager.pool_category == pool_category)
        if primary_strategy:
            query = query.filter(Manager.primary_strategy == primary_strategy)
        managers = query.limit(20).all()
        total = query.count()
        rows = []
        for m in managers:
            pc = self.db.query(func.count(Product.id)).filter(Product.manager_id == m.id).scalar()
            rows.append({
                "管理人名称": m.manager_name,
                "策略": strategy_map.get(m.primary_strategy, m.primary_strategy or "-"),
                "跟踪池": pool_map.get(m.pool_category, m.pool_category or "-"),
                "规模": m.aum_range or "-",
                "产品数": pc,
            })
        return {
            "columns": ["管理人名称", "策略", "跟踪池", "规模", "产品数"],
            "rows": rows, "total": total
        }

    def _preview_product(self, **kwargs) -> dict:
        strategy_type = kwargs.get('strategy_type')
        query = self.db.query(Product)
        if strategy_type:
            query = query.filter(Product.strategy_type == strategy_type)
        products = query.limit(20).all()
        total = query.count()
        rows = []
        for p in products:
            mgr = self.db.query(Manager.manager_name).filter(Manager.id == p.manager_id).scalar() if p.manager_id else "-"
            latest = self.db.query(NavData).filter(NavData.product_id == p.id).order_by(NavData.nav_date.desc()).first()
            rows.append({
                "产品代码": p.product_code or "-",
                "产品名称": p.product_name,
                "管理人": mgr or "-",
                "策略类型": p.strategy_type or "-",
                "最新净值": f"{latest.unit_nav:.4f}" if latest else "-",
                "净值日期": str(latest.nav_date) if latest else "-",
            })
        return {
            "columns": ["产品代码", "产品名称", "管理人", "策略类型", "最新净值", "净值日期"],
            "rows": rows, "total": total
        }

    def _preview_portfolio(self, **kwargs) -> dict:
        portfolio_id = kwargs.get('portfolio_id')
        query = self.db.query(Portfolio)
        if portfolio_id:
            query = query.filter(Portfolio.id == portfolio_id)
        portfolios = query.limit(20).all()
        total = query.count()
        rows = []
        for p in portfolios:
            comp_count = self.db.query(func.count(PortfolioComponent.id)).filter(
                PortfolioComponent.portfolio_id == p.id
            ).scalar()
            rows.append({
                "组合名称": p.name,
                "描述": p.description or "-",
                "成分数": comp_count,
                "创建时间": str(p.created_at.date()) if p.created_at else "-",
            })
        return {
            "columns": ["组合名称", "描述", "成分数", "创建时间"],
            "rows": rows, "total": total
        }

    def _preview_project(self, **kwargs) -> dict:
        stage = kwargs.get('stage')
        stage_map = {
            "sourcing": "Sourcing", "initial_screening": "初筛", "due_diligence": "尽调",
            "investment_decision": "投决", "post_investment": "投后", "exit": "退出"
        }
        query = self.db.query(Project)
        if stage:
            query = query.filter(Project.stage == stage)
        projects = query.limit(20).all()
        total = query.count()
        rows = []
        for p in projects:
            rows.append({
                "项目名称": p.project_name,
                "行业": p.industry or "-",
                "阶段": stage_map.get(p.stage, p.stage or "-"),
                "投资额(万)": str(p.investment_amount) if p.investment_amount else "-",
                "估值(万)": str(p.valuation) if p.valuation else "-",
            })
        return {
            "columns": ["项目名称", "行业", "阶段", "投资额(万)", "估值(万)"],
            "rows": rows, "total": total
        }

    # ========== 净值数据导出 ==========
    
    def export_nav_data(
        self,
        product_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> io.BytesIO:
        """导出产品净值数据"""
        product = self.db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise ValueError(f"产品不存在: {product_id}")
        
        query = self.db.query(NavData).filter(NavData.product_id == product_id)
        if start_date:
            query = query.filter(NavData.nav_date >= start_date)
        if end_date:
            query = query.filter(NavData.nav_date <= end_date)
        
        nav_list = query.order_by(NavData.nav_date.asc()).all()
        
        data = []
        for nav in nav_list:
            data.append({
                "产品代码": product.product_code or "",
                "产品名称": product.product_name,
                "净值日期": str(nav.nav_date),
                "单位净值": nav.unit_nav,
                "累计净值": nav.cumulative_nav or "",
            })
        
        wb = Workbook()
        ws = wb.active
        ws.title = "净值数据"
        
        df = pd.DataFrame(data)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_style(ws, 1, len(df.columns))
        self._auto_column_width(ws)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
